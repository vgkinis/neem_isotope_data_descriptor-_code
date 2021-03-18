import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, FixedLocator, FixedFormatter
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib import dates as mpl_dates
import string
import time
from datetime import datetime
import copy
import os
import subprocess
import sys
sys.path.append("/Users/vasilis")
import vaspy
from vaspy import syttensen
plt.rc('text', usetex=True)
plt.rc('font', family='sans-serif', size = 16)
plt.rc('text.latex', preamble=r'\usepackage{cmbright}')
# plt.rc('font', size = 12)
plt.rc('xtick', direction = 'in')
plt.rc('ytick', direction = 'in')
mylabelsize = 16
plt.rc('ytick', labelsize = mylabelsize)
plt.rc('xtick', labelsize = mylabelsize)
plt.rc('axes', labelsize = mylabelsize)

plt.close("all")
plt.ion()


def return_unique_runs(instrument = None):
    """
    Reads the neem_nd_all file and finds the unique hits of run ID's included in it
    if instrument param is not None then it returns only those run Id's
    with the instrument string contained in ID

    instrument should be 'HBDS', 'HIDS' or 'HKDS'
    """
    summary = np.genfromtxt("../neem_nd_all.txt", delimiter = "\t", dtype = \
    [('run_id', "U60"), ('bag', int), ('sample', int), ('depth', float), ('d18', float), \
    ('dD', float)], skip_header = 1)
    unique_runs = pd.Series(np.unique(summary['run_id']))
    print(unique_runs)
    if instrument:
        return unique_runs[unique_runs.str.contains(instrument)].reset_index(drop = True)
    else:
        return unique_runs





def correct_standard_values_openpyxl(runs_series, import_folder = "./all_neem_xls_runs", export_folder = "all_neem_xls_runs_corr"):
    """

    """
    flog = open("./correct_standard_values_openpyxl.log", "w")
    counter = 0
    for run in runs_series:
        print("opening file %s" %run)

        filein = "./all_neem_xls_runs/" + run + ".xlsx"
        wb = openpyxl.load_workbook(filename = filein)
        print(wb.sheetnames)
        ws_calibrated_data = wb[wb.sheetnames[1]]
        run_standards = {}
        run_standards["std1_real_d18"] = ws_calibrated_data["N5"].value
        run_standards["std1_real_dD"] = ws_calibrated_data["O5"].value
        run_standards["std2_real_d18"] = ws_calibrated_data["N4"].value
        run_standards["std2_real_dD"] = ws_calibrated_data["O4"].value
        run_standards["std3_real_d18"] = ws_calibrated_data["N6"].value
        run_standards["std3_real_dD"] = ws_calibrated_data["O6"].value
        print(run_standards)

        if run_standards["std2_real_d18"] > -35:
            print("Run with IG standards")
            ws_calibrated_data["N4"] = -33.5
            ws_calibrated_data["O4"] = -257.45
            ws_calibrated_data["N5"] = -21.88
            ws_calibrated_data["O5"] = -168.45
            ws_calibrated_data["N6"] = -39.98
            ws_calibrated_data["O6"] = -311.11
        elif run_standards["std2_real_d18"] < -36:
            print("Run with GL standards")
            ws_calibrated_data["N4"] = -39.98
            ws_calibrated_data["O4"] = -311.11
            ws_calibrated_data["N5"] = -33.5
            ws_calibrated_data["O5"] = -257.45
            ws_calibrated_data["N6"] = -54.07
            ws_calibrated_data["O6"] = -428.2

        outpath = "./all_neem_xls_runs_corr/" + run + "_corr.xlsx"
        wb.save(outpath)


    os.chdir("./all_neem_xls_runs_corr")
    os.system("/Applications/LibreOffice.app/Contents/MacOS/soffice  --headless --convert-to xlsx --outdir ./all_neem_xls_runs_corr *.xlsx")
    os.chdir("..")
    flog.close()
    return



def runs_build_data_frame_corr(runs_series, printout_df = True, save_csv = True):
    """
    receives a pandas Series with the xls filenames runs
    for every run it reads the xlls file (both sheets in xls)
    it builds a stats dictionary with the data of interest
    it converts the dictionary to a data frame
    it appends the new data frame to a big data frame called stats_df_main
    keys of stats_df_main are:

    Index([u'mean_std2_dD', u'offset_d18', u'top_depth', u'ok', u'N2_flag',
       u'h2o_mean', u'offset_dD', u'h2o_std', u'slope_dD', u'insters_d18',
       u'inters_dD', u'slope_d18', u'std_std2_d18', u'mean_std2_d18',
       u'std_std2_dD'],
      dtype='object')

      And the dataframe is indexed by the time stamp of each xls sheet
    """

    flog = open("./runs_build_data_frame_corr.log", "w")
    counter = 0
    for run in runs_series:
        run = run + "_corr"
        print("opening file %s" %run)
        try:
            data_one_xls = pd.read_excel("./all_neem_xls_runs_corr_stds/" + run + ".xlsx", sheet_name = None, na_values = ["", " ", 14*" "], header = None)
        except:
            print("Problem opening run %s" %run)
            flog.write("Problem opening run %s\n" %run)
            continue
        try:
            raw_data = data_one_xls["Paste Raw Data Here"]
            cal_data = data_one_xls["Calibrated data"]
        except:
            raw_data = data_one_xls = pd.read_excel("./all_neem_xls_runs_corr_stds/" + run + ".xlsx", sheet_name = 0, na_values = ["", " ", 14*" "], verbose = True, header = None)
            cal_data = data_one_xls = pd.read_excel("./all_neem_xls_runs_corr_stds/" + run + ".xlsx", sheet_name = 1, na_values = ["", " ", 14*" "], verbose = True, header = None)

        stats = {}
        stats["run_ID"] = run
        stats["offset_d18"] = cal_data.iloc[3,17]
        stats["offset_dD"] = cal_data.iloc[3,18]
        stats["ok"] = cal_data.iloc[6,16]
        stats["mean_std2_d18"] = cal_data.iloc[16,12]
        stats["mean_std2_dD"] = cal_data.iloc[16,13]
        stats["std_std2_d18"] = cal_data.iloc[17,12]
        stats["std_std2_dD"] = cal_data.iloc[17,13]
        stats["mean_std2_d18_raw"] = cal_data.iloc[3,11]
        stats["mean_std2_dD_raw"] = cal_data.iloc[3,12]
        stats["slope_d18"] = cal_data.iloc[8,12]
        stats["slope_dD"] = cal_data.iloc[8,13]
        stats["inters_d18"] = cal_data.iloc[9,12]
        stats["inters_dD"] = cal_data.iloc[9,13]

        if np.isnan(stats["mean_std2_d18"]):
            stats["mean_std2_d18"] = cal_data.iloc[18,12]
            stats["mean_std2_dD"] = cal_data.iloc[18,13]
            stats["std_std2_d18"] = cal_data.iloc[19,12]
            stats["std_std2_dD"] = cal_data.iloc[19,13]
        try:
            stats["N2_flag"] = raw_data["n2_flag"][1]
        except:
            stats["N2_flag"] = 0

        try:
            stats["top_depth"] = np.round(cal_data.iloc[0,20])*0.55-0.55
            h2o = pd.Series(cal_data.iloc[:,4])
            stats["h2o_mean"] = np.mean(h2o[10:])
            stats["h2o_std"] = np.std(h2o[10:])
            stats["time_stamp"] = raw_data.iloc[1,2].strip()
        except Exception as e:
            flog.write("Problem with run %s: %s\n" %(run, e))
            continue

        stats_df = pd.DataFrame.from_dict(stats, orient = 'index').T
        stats_df.set_index(keys = 'time_stamp', inplace = True)
        print(stats_df.to_string())
        if counter !=0:
            stats_df_main = stats_df_main.append(stats_df)
        elif counter == 0:
            stats_df_main = copy.deepcopy(stats_df)
        counter+=1

    if printout_df == True:
        print("\n\n\n\n")
        print(stats_df_main.to_string())

    stats_df_main_sorted = stats_df_main.reindex(columns = sorted(stats_df_main.columns), copy = True)

    if save_csv:
        stats_df_main_sorted.to_csv("./stats_df_main_corr.csv", sep = "\t", na_rep = "NaN")
    flog.close()
    return stats_df_main_sorted



###########################################################################
#  Plotting routines
###########################################################################



def plot_offset_std2_vs_depth_vs_time():
    """

    """
    years = mpl_dates.YearLocator()
    f0, axess = plt.subplots(nrows = 2, ncols = 2, num = 200, figsize = (10,6), tight_layout = True)
    ax0 = axess[0][0]
    ax0.set_ylabel(r'$\delta^{18}\mathrm{O}$ offset [\textperthousand]')
    ax0.set_xlabel(r'Measurement Time')
    ax0.xaxis.set_major_locator(years)

    ax1 = axess[0][1]
    ax1.set_ylabel(r'$\delta\mathrm{D}$ offset [\textperthousand]')
    ax1.set_xlabel(R"Measurement Time")
    ax1.xaxis.set_major_locator(years)

    ax2 = axess[1][0]
    ax2.set_ylabel(r'$\delta^{18}\mathrm{O}$ offset [\textperthousand]')
    ax2.set_xlabel(r'Depth [m]')

    ax3 = axess[1][1]
    ax3.set_ylabel(r'$\delta\mathrm{D}$ offset [\textperthousand]')
    ax3.set_xlabel(r"Depth [m]")

    ##############################
    #  Read xls data
    ##############################

    all_stats = pd.read_excel("./stats_all_neem_runs.xlsx", sheet_name = None, na_values = ["Nan"])
    df_HBDS = all_stats["stats_NEEM_HBDS_corr"].set_index(keys = 'time_stamp')
    df_HBDS = df_HBDS[df_HBDS["top_depth"]>1210.55]
    df_HIDS = all_stats["stats_NEEM_HIDS_corr"].set_index(keys = 'time_stamp')
    df_HIDS = df_HIDS[df_HIDS["top_depth"]>1210.55]
    df_HKDS = all_stats["stats_NEEM_HKDS_corr"].set_index(keys = 'time_stamp')
    df_HKDS = df_HKDS[df_HKDS["top_depth"]>1210.55]


    ax0.scatter(x = pd.to_datetime(df_HBDS.index), y = df_HBDS['offset_d18'], s = 2, c = "g")
    ax0.scatter(x = pd.to_datetime(df_HIDS.index), y = df_HIDS['offset_d18'], s = 2, c = "b")
    ax0.scatter(x = pd.to_datetime(df_HKDS.index), y = df_HKDS['offset_d18'], s = 2, c = "r")

    ax1.scatter(x = pd.to_datetime(df_HBDS.index), y = df_HBDS['offset_dD'], s = 2, c = "g")
    ax1.scatter(x = pd.to_datetime(df_HIDS.index), y = df_HIDS['offset_dD'], s = 2, c = "b")
    ax1.scatter(x = pd.to_datetime(df_HKDS.index), y = df_HKDS['offset_dD'], s = 2, c = "r")

    ax2.scatter(x = df_HBDS['top_depth'], y = df_HBDS['offset_d18'], s = 2, c = "g")
    ax2.scatter(x = df_HIDS['top_depth'], y = df_HIDS['offset_d18'], s = 2, c = "b")
    ax2.scatter(x = df_HKDS['top_depth'], y = df_HKDS['offset_d18'], s = 2, c = "r")

    ax3.scatter(x = df_HBDS['top_depth'], y = df_HBDS['offset_dD'], s = 2, c = "g")
    ax3.scatter(x = df_HIDS['top_depth'], y = df_HIDS['offset_dD'], s = 2, c = "b")
    ax3.scatter(x = df_HKDS['top_depth'], y = df_HKDS['offset_dD'], s = 2, c = "r")


    return


def plot_std_std2_vs_depth_vs_time():
    """

    """
    years = mpl_dates.YearLocator()
    f0, axess = plt.subplots(nrows = 2, ncols = 2, num = 312, figsize = (10,6), tight_layout = True)
    ax0 = axess[0][0]
    ax0.set_ylabel(r'$\delta^{18}\mathrm{O}\,\,1\sigma$ [\textperthousand]')
    ax0.set_xlabel(r'Measurement Time')
    ax0.xaxis.set_major_locator(years)

    ax1 = axess[0][1]
    ax1.set_ylabel(r'$\delta\mathrm{D}\,\,1\sigma$ [\textperthousand]')
    ax1.set_xlabel(R"Measurement Time")
    ax1.xaxis.set_major_locator(years)

    ax2 = axess[1][0]
    ax2.set_ylabel(r'$\delta^{18}\mathrm{O}\,\,1\sigma$ [\textperthousand]')
    ax2.set_xlabel(r'Depth [m]')

    ax3 = axess[1][1]
    ax3.set_ylabel(r'$\delta\mathrm{D}\,\,1\sigma$ [\textperthousand]')
    ax3.set_xlabel(r"Depth [m]")

    ##############################
    #  Read xls data
    ##############################

    all_stats = pd.read_excel("./stats_all_neem_runs.xlsx", sheet_name = None, na_values = ["Nan"])
    df_HBDS = all_stats["stats_NEEM_HBDS_corr"].set_index(keys = 'time_stamp')
    df_HBDS = df_HBDS[df_HBDS["top_depth"]>1210.55]
    df_HIDS = all_stats["stats_NEEM_HIDS_corr"].set_index(keys = 'time_stamp')
    df_HIDS = df_HIDS[df_HIDS["top_depth"]>1210.55]
    df_HKDS = all_stats["stats_NEEM_HKDS_corr"].set_index(keys = 'time_stamp')
    df_HKDS = df_HKDS[df_HKDS["top_depth"]>1210.55]


    ax0.scatter(x = pd.to_datetime(df_HBDS.index), y = df_HBDS['std_std2_d18'], s = 2, c = "g")
    ax0.scatter(x = pd.to_datetime(df_HIDS.index), y = df_HIDS['std_std2_d18'], s = 2, c = "b")
    ax0.scatter(x = pd.to_datetime(df_HKDS.index), y = df_HKDS['std_std2_d18'], s = 2, c = "r")

    ax1.scatter(x = pd.to_datetime(df_HBDS.index), y = df_HBDS['std_std2_dD'], s = 2, c = "g")
    ax1.scatter(x = pd.to_datetime(df_HIDS.index), y = df_HIDS['std_std2_dD'], s = 2, c = "b")
    ax1.scatter(x = pd.to_datetime(df_HKDS.index), y = df_HKDS['std_std2_dD'], s = 2, c = "r")

    ax2.scatter(x = df_HBDS['top_depth'], y = df_HBDS['std_std2_d18'], s = 2, c = "g")
    ax2.scatter(x = df_HIDS['top_depth'], y = df_HIDS['std_std2_d18'], s = 2, c = "b")
    ax2.scatter(x = df_HKDS['top_depth'], y = df_HKDS['std_std2_d18'], s = 2, c = "r")

    ax3.scatter(x = df_HBDS['top_depth'], y = df_HBDS['std_std2_dD'], s = 2, c = "g")
    ax3.scatter(x = df_HIDS['top_depth'], y = df_HIDS['std_std2_dD'], s = 2, c = "b")
    ax3.scatter(x = df_HKDS['top_depth'], y = df_HKDS['std_std2_dD'], s = 2, c = "r")


    return



def plot_offset_std_std2_vs_depth():
    """

    """
    years = mpl_dates.YearLocator()
    f0, axess = plt.subplots(nrows = 2, ncols = 2, num = 9400, figsize = (10,6), tight_layout = True)
    ax0 = axess[0][0]
    ax0.set_ylabel(r'$\delta^{18}\mathrm{O}\,\,1-\sigma$ [\textperthousand]')
    ax0.set_xlabel(r'Depth [m]')
    ax0.set_xticks((1250, 1500, 1750, 2000, 2250, 2500))


    ax1 = axess[0][1]
    ax1.set_ylabel(r'$\delta\mathrm{D}\,\,1-\sigma$ [\textperthousand]')
    ax1.set_xlabel(r"Depth [m]")
    ax1.set_xticks((1250, 1500, 1750, 2000, 2250, 2500))


    ax2 = axess[1][0]
    ax2.set_ylabel(r'$\delta^{18}\mathrm{O}$ offset [\textperthousand]')
    ax2.set_xlabel(r'Depth [m]')
    ax2.set_xticks((1250, 1500, 1750, 2000, 2250, 2500))

    ax3 = axess[1][1]
    ax3.set_ylabel(r'$\delta\mathrm{D}$ offset [\textperthousand]')
    ax3.set_xlabel(r"Depth [m]")
    ax3.set_xticks((1250, 1500, 1750, 2000, 2250, 2500))

    ##############################
    #  Read xls data
    ##############################

    all_stats = pd.read_excel("./stats_all_neem_runs.xlsx", sheet_name = None, na_values = ["Nan"])
    df_HBDS = all_stats["stats_NEEM_HBDS_corr"].set_index(keys = 'time_stamp')
    df_HBDS = df_HBDS[df_HBDS["top_depth"]>1210.55]
    df_HIDS = all_stats["stats_NEEM_HIDS_corr"].set_index(keys = 'time_stamp')
    df_HIDS = df_HIDS[df_HIDS["top_depth"]>1210.55]
    df_HKDS = all_stats["stats_NEEM_HKDS_corr"].set_index(keys = 'time_stamp')
    df_HKDS = df_HKDS[df_HKDS["top_depth"]>1210.55]


    ax0.scatter(x = df_HBDS['top_depth'], y = df_HBDS['std_std2_d18'], s = 2, c = "g", label = "L2120-i")
    ax0.scatter(x = df_HIDS['top_depth'], y = df_HIDS['std_std2_d18'], s = 2, c = "b", label = "L2130-i")
    ax0.scatter(x = df_HKDS['top_depth'], y = df_HKDS['std_std2_d18'], s = 2, c = "r", label = "L2140-i")
    ax0.legend(markerscale = 2, frameon = False, loc = 2)

    ax1.scatter(x = df_HBDS['top_depth'], y = df_HBDS['std_std2_dD'], s = 2, c = "g")
    ax1.scatter(x = df_HIDS['top_depth'], y = df_HIDS['std_std2_dD'], s = 2, c = "b")
    ax1.scatter(x = df_HKDS['top_depth'], y = df_HKDS['std_std2_dD'], s = 2, c = "r")

    ax2.scatter(x = df_HBDS['top_depth'], y = df_HBDS['offset_d18'], s = 2, c = "g")
    ax2.scatter(x = df_HIDS['top_depth'], y = df_HIDS['offset_d18'], s = 2, c = "b")
    ax2.scatter(x = df_HKDS['top_depth'], y = df_HKDS['offset_d18'], s = 2, c = "r")

    ax3.scatter(x = df_HBDS['top_depth'], y = df_HBDS['offset_dD'], s = 2, c = "g")
    ax3.scatter(x = df_HIDS['top_depth'], y = df_HIDS['offset_dD'], s = 2, c = "b")
    ax3.scatter(x = df_HKDS['top_depth'], y = df_HKDS['offset_dD'], s = 2, c = "r")


    return




def plot_offset_std_std2_histograms():
    """

    """
    nbins = 12
    years = mpl_dates.YearLocator()
    f0, axess = plt.subplots(nrows = 2, ncols = 2, num = 302, figsize = (8,8), tight_layout = True)
    ax0 = axess[0][0]
    ax0.set_xlabel(r'$\delta^{18}\mathrm{O}\,\,1\sigma$ [\textperthousand]', size = 16)
    ax0.set_xlim((0, 0.15))
    ax0.set_yticks((5, 10, 15, 20, 25, 30))
    # ax0.set_xticklabels(("0", "0.025", "0.50", "0.075", "0.1", "0.125", "0.15"))

    ax1 = axess[0][1]
    ax1.set_xlabel(r'$\delta\mathrm{D}\,\,1\sigma$ [\textperthousand]')
    ax1.set_xlim(0,1)
    ax1.set_yticks((1, 2, 3, 4))

    ax2 = axess[1][0]
    ax2.set_xlabel(r'$\delta^{18}\mathrm{O}$ offset [\textperthousand]')
    ax2.set_yticks((2, 4, 6, 8, 10, 12))

    ax3 = axess[1][1]
    ax3.set_xlabel(r'$\delta\mathrm{D}$ offset [\textperthousand]')
    ax3.set_yticks((1, 2, 3, 4))

    ##############################
    #  Read xls data
    ##############################

    all_stats = pd.read_excel("./stats_all_neem_runs.xlsx", sheet_name = None, na_values = ["Nan"])
    df_HBDS = all_stats["stats_NEEM_HBDS_corr"].set_index(keys = 'time_stamp')
    df_HBDS = df_HBDS[df_HBDS["top_depth"]>1210.55]
    df_HIDS = all_stats["stats_NEEM_HIDS_corr"].set_index(keys = 'time_stamp')
    df_HIDS = df_HIDS[df_HIDS["top_depth"]>1210.55]
    df_HKDS = all_stats["stats_NEEM_HKDS_corr"].set_index(keys = 'time_stamp')
    df_HKDS = df_HKDS[df_HKDS["top_depth"]>1210.55]

    ax0.hist(df_HBDS['std_std2_d18'], nbins, density = True, histtype = 'step', color = 'g', linewidth = 1, label = "L2120-i")
    ax0.hist(df_HIDS['std_std2_d18'], nbins, density = True, histtype = 'step', color = 'b', linewidth = 1, label = "L2130-i")
    ax0.hist(df_HKDS['std_std2_d18'], nbins, density = True, histtype = 'step', color = 'r', linewidth = 1, label = "L2140-i")
    ax0.legend(frameon = False)


    ax1.hist(df_HBDS['std_std2_dD'], nbins, density = True, histtype = 'step', color = 'g', linewidth = 1)
    ax1.hist(df_HIDS['std_std2_dD'], nbins, density = True, histtype = 'step', color = 'b', linewidth = 1)
    ax1.hist(df_HKDS['std_std2_dD'], nbins, density = True, histtype = 'step', color = 'r', linewidth = 1)
    # ax1.axvline(x = np.mean(df_HBDS['std_std2_dD']), ymin = 0, ymax = 1, linestyle = "-.", color = "g", linewidth = 1.)
    # ax1.axvline(x = np.mean(df_HIDS['std_std2_dD']), ymin = 0, ymax = 1, linestyle = "-.", color = "b", linewidth = 1.)
    # ax1.axvline(x = np.mean(df_HKDS['std_std2_dD']), ymin = 0, ymax = 1, linestyle = "-.", color = "r", linewidth = 1.)

    ax2.hist(df_HBDS['offset_d18'], nbins, density = True, histtype = 'step', color = 'g', linewidth = 1)
    ax2.hist(df_HIDS['offset_d18'], nbins, density = True, histtype = 'step', color = 'b', linewidth = 1)
    ax2.hist(df_HKDS['offset_d18'], nbins, density = True, histtype = 'step', color = 'r', linewidth = 1)


    ax3.hist(df_HBDS['offset_dD'], nbins, density = True, histtype = 'step', color = 'g', linewidth = 1)
    ax3.hist(df_HIDS['offset_dD'], nbins, density = True, histtype = 'step', color = 'b', linewidth = 1)
    ax3.hist(df_HKDS['offset_dD'], nbins, density = True, histtype = 'step', color = 'r', linewidth = 1)

    print("HBDS\tHIDS\tHKDS")
    print("%0.3f\t%0.3f\t%0.3f" %(np.mean(df_HBDS["std_std2_d18"]), np.mean(df_HIDS["std_std2_d18"]), np.mean(df_HKDS["std_std2_d18"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.std(df_HBDS["std_std2_d18"]), np.std(df_HIDS["std_std2_d18"]), np.std(df_HKDS["std_std2_d18"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.mean(df_HBDS["std_std2_dD"]), np.mean(df_HIDS["std_std2_dD"]), np.mean(df_HKDS["std_std2_dD"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.std(df_HBDS["std_std2_dD"]), np.std(df_HIDS["std_std2_dD"]), np.std(df_HKDS["std_std2_dD"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.mean(df_HBDS["offset_d18"]), np.mean(df_HIDS["offset_d18"]), np.mean(df_HKDS["offset_d18"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.std(df_HBDS["offset_d18"]), np.std(df_HIDS["offset_d18"]), np.std(df_HKDS["offset_d18"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.mean(df_HBDS["offset_dD"]), np.mean(df_HIDS["offset_dD"]), np.mean(df_HKDS["offset_dD"])))
    print("%0.3f\t%0.3f\t%0.3f" %(np.std(df_HBDS["offset_dD"]), np.std(df_HIDS["offset_dD"]), np.std(df_HKDS["offset_dD"])))

    return

def plot_time_vs_depth():
    """

    """
    years = mpl_dates.YearLocator()
    f0, axess = plt.subplots(nrows = 1, ncols = 1, num = 400, figsize = (6,6), tight_layout = True)
    ax0 = axess
    ax0.set_ylabel(r'Depth [m]')
    ax0.set_xlabel(r'Measurement Time')
    ax0.xaxis.set_major_locator(years)

    ##############################
    #  Read xls data
    ##############################

    all_stats = pd.read_excel("./stats_all_neem_runs.xlsx", sheet_name = None, na_values = ["Nan"])
    df_HBDS = all_stats["stats_NEEM_HBDS_corr"].set_index(keys = 'time_stamp')
    df_HBDS = df_HBDS[df_HBDS["top_depth"]>1210.55]
    df_HIDS = all_stats["stats_NEEM_HIDS_corr"].set_index(keys = 'time_stamp')
    df_HIDS = df_HIDS[df_HIDS["top_depth"]>1210.55]
    df_HKDS = all_stats["stats_NEEM_HKDS_corr"].set_index(keys = 'time_stamp')
    df_HKDS = df_HKDS[df_HKDS["top_depth"]>1210.55]

    ax0.scatter(x = pd.to_datetime(df_HBDS.index), y = df_HBDS['top_depth'], s = 2, c = "g", label = "L2120-i")
    ax0.scatter(x = pd.to_datetime(df_HIDS.index), y = df_HIDS['top_depth'], s = 2, c = "b", label = "L2130-i")
    ax0.scatter(x = pd.to_datetime(df_HKDS.index), y = df_HKDS['top_depth'], s = 2, c = "r", label = "L2140-i")
    ax0.legend(markerscale = 4, frameon = False)

    return



def plot_nd_all_vs_instrument():
    """

    """
    instruments = ['HBDS', 'HKDS', 'HIDS', 'HHHS']
    nd_all_df = pd.read_csv("../neem_nd_all_corr.txt", delimiter = "\t")
    nd_all_df = nd_all_df[nd_all_df["Depth"]>1210.55]
    print(np.shape(nd_all_df))

    f0, axess = plt.subplots(nrows = 1, ncols = 1, num = 500, figsize = (10,6), tight_layout = True)
    ax0 = axess
    ax0.set_xlabel(r'Depth [m]')
    ax0.set_ylabel(r'$\delta^{18}\mathrm{O}$ [\textperthousand]')

    nd_all_instrument_df = nd_all_df[nd_all_df["Run_ID"].str.contains('HBDS')].reset_index(drop = True)
    ax0.scatter(x = nd_all_instrument_df["Depth"], y = nd_all_instrument_df["d18"], s = 1, c = 'g', label = "L2120-i")
    print(np.shape(nd_all_instrument_df))

    nd_all_instrument_df = nd_all_df[nd_all_df["Run_ID"].str.contains('HIDS')].reset_index(drop = True)
    ax0.scatter(x = nd_all_instrument_df["Depth"], y = nd_all_instrument_df["d18"], s = 1, c = 'b', label = "L2130-i")
    print(np.shape(nd_all_instrument_df))

    nd_all_instrument_df = nd_all_df[nd_all_df["Run_ID"].str.contains('HKDS')].reset_index(drop = True)
    ax0.scatter(x = nd_all_instrument_df["Depth"], y = nd_all_instrument_df["d18"], s = 1, c = 'r', label = "L2140-i")
    print(np.shape(nd_all_instrument_df))

    nd_all_instrument_df = nd_all_df[nd_all_df["Run_ID"].str.contains('HHHS')].reset_index(drop = True)
    ax0.scatter(x = nd_all_instrument_df["Depth"], y = nd_all_instrument_df["d18"], s = 1, c = 'k', label = "interpolation")
    print(np.shape(nd_all_instrument_df))

    ax0.legend(markerscale = 6, frameon = False, loc = 3)

    return


def plot_neem_nd_all():
    # data_pre = np.genfromtxt("../neem_nd_all.txt", delimiter = "\t", dtype = \
    # [('run_id', "U60"), ('bag', int), ('sample', int), ('depth', float), ('d18', float), \
    # ('dD', float)], skip_header = 1)

    data_corr = np.genfromtxt("../neem_nd_all_corr.txt", delimiter = "\t", dtype = \
    [('run_id', "U60"), ('bag', int), ('sample', int), ('depth', float), ('d18', float), \
    ('dD', float)], skip_header = 1)
    data_corr = data_corr[data_corr["depth"]>1210.55]

    f0, axess = plt.subplots(nrows = 1, ncols = 1 , num = 510, figsize = (10,6), tight_layout = True)
    f0.suptitle("NEEM 0.05 m resolution record", size = 18)
    ax0 = axess
    ax0.plot(data_corr["depth"], data_corr["dD"], linewidth = 1, color = "r")
    ax0.set_ylabel(r'$\delta\mathrm{D}$ [\textperthousand]')
    ax0.yaxis.set_label_coords(-0.08,0.33)
    ax0.set_xlabel(r'Depth [m]')
    ax0.set_ylim((-400, -100))
    ax0.spines['left'].set_bounds(-400, -210)
    ax0.set_yticks((-350, -300, -250))
    ax0.spines['top'].set_visible(False)
    ax0.spines['right'].set_visible(False)


    ax00 = ax0.twinx()
    ax00.plot(data_corr["depth"], data_corr["d18"], linewidth = 1, color = "b")
    ax00.set_ylabel(r'${\delta}^{18}\mathrm{O}$ [\textperthousand]', rotation = 270)
    ax00.yaxis.set_label_coords(+1.09,0.72)
    ax00.set_xlabel(r'Depth [m]')
    ax00.set_ylim((-72,-28))
    ax00.spines['right'].set_bounds(-52,-28)
    ax00.set_yticks((-50, -40, -30))
    ax00.spines['top'].set_visible(False)
    ax00.spines['left'].set_visible(False)





    f1, axess1 = plt.subplots(nrows = 1, ncols = 1 , num = 802, figsize = (10,6), tight_layout = True)
    ax1 = axess1
    ax1.plot(data_corr["depth"], data_corr["d18"], linewidth = 1, color = "b")
    ax1.set_ylabel(r'$\delta^{18}\mathrm{O}$ [\textperthousand]')
    ax1.set_xlabel(r'Depth [m]')

    f2, axess2 = plt.subplots(nrows = 1, ncols = 1 , num = 579, figsize = (10,6), tight_layout = True)
    ax2 = axess2
    ax2.plot(data_corr["depth"], data_corr["dD"] - 8*data_corr["d18"], linewidth = 1, color = "g")
    ax2.set_ylabel(r'$\mathrm{D}_\mathrm{xs}$ [\textperthousand]')
    ax2.set_xlabel(r'Depth [m]')


    return


def plot_injection_pulse():
    """

    """
    files_in_dir = os.listdir("./10")
    counter = 0
    for fil in files_in_dir:
        if fil[:2] == "HI":
            bunch10 = syttensen.bunch_syttensen.Bunch()
            bunch10.read_data("./10/%s" %fil)
        if counter == 0:
            big_bunch = bunch10
            counter+=1
            continue
        else:
            big_bunch.concat(bunch10)
        counter+=1
    big_bunch.plot()
    print(big_bunch.epoch)
    print((big_bunch.epoch[-1] - big_bunch.epoch[0])/3600)
    datetime_obj = []
    for j in np.arange(np.size(big_bunch.epoch)):
        datetime_obj_j = datetime.fromtimestamp(big_bunch.epoch[j])
        datetime_obj.append(datetime_obj_j)
    big_bunch.date_time = np.array((datetime_obj))
    print(big_bunch.date_time)

    f0, axess = plt.subplots(nrows = 1, ncols = 1 , num = 389, figsize = (6,5), tight_layout = True)
    axess1 = axess.twinx()
    # axess.plot_date(big_bunch.date_time, big_bunch.h2o, linewidth  = 1, fmt = "-", color = "#332e2c")
    # axess1.plot_date(big_bunch.date_time, big_bunch.dD, linewidth = 1, fmt = "-", color = "#db724f")
    # date_format = mpl_dates.DateFormatter('%H:%M')
    # axess.xaxis.set_major_formatter(date_format)
    axess1.plot(big_bunch.secs, big_bunch.h2o, linewidth  = 1, color = "#332e2c")
    axess.plot(big_bunch.secs, big_bunch.dD, linewidth = 1, color = "#d6472b")
    axess.set_xlabel("Run Time [s]")
    axess.set_ylabel(r'$\delta\mathrm{D}$ [\textperthousand]', color = '#d6472b')
    axess1.set_ylabel(r'$[{\mathrm{H}}_{2}\mathrm{O}] [\mathrm{ppm}]$', color = '#332e2c', rotation = 270, labelpad = 20)
    axess.set_ylim((-600, 0))
    axess.spines["left"].set_color('#d6472b')
    axess1.spines["left"].set_color('#d6472b')
    axess1.spines["right"].set_color('#332e2c')
    axess.spines["right"].set_color('#332e2c')
    axess.tick_params(axis = 'y', color = '#d6472b', labelcolor = '#d6472b')
    axess1.tick_params(axis = 'y', color = '#332e2c', labelcolor = '#332e2c')



    return f0, axess, axess1


def plot_cordinator_run():
    """

    """
    f = open("./HIDS2159_Fast_IsoWater_20160310_125444_corr.txt", "r")
    data_in = np.genfromtxt(f, skip_header = 1, usecols = (0,2,3,5,6,7), \
    dtype = None, names = ["nr", "dtime_string", "tray_pos", "d18", "dD", "h2o"], encoding = "None", delimiter = "\t", autostrip = True)
    print(data_in["dD"])
    print(data_in)
    dtime_obj = []
    for j in np.arange(np.size(data_in["dtime_string"])):
        dtime_obj.append(datetime.strptime(data_in["dtime_string"][j], "%Y/%m/%d %H:%M:%S"))
        print(type(dtime_obj[j]))
        data_in["tray_pos"][j] = np.int(data_in["tray_pos"][j][-2:])

    f0, axess = plt.subplots(nrows = 1, ncols = 1 , num = 567, figsize = (6,5), tight_layout = True)
    axess.plot_date(dtime_obj, data_in["dD"], linewidth  = 1.2, fmt = "-", color = "#d6472b")
    axess1 = axess.twinx()
    axess1.plot_date(dtime_obj, data_in["tray_pos"], linewidth = 0.78, fmt = "-", color = "#787675")
    date_format = mpl_dates.DateFormatter('%H:%M')
    axess.xaxis.set_major_formatter(date_format)
    axess1.set_yticks([0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55])
    axess.set_xlabel("Time")
    axess.set_ylabel(r"$\delta\mathrm{D}$ [\textperthousand]", color = '#d6482b')
    axess1.set_ylabel("Tray position", rotation = 270, labelpad = 20)
    axess.tick_params(axis = 'y', color = '#d6472b', labelcolor = '#d6472b')
    axess.spines["left"].set_color('#d6472b')
    axess1.spines["left"].set_color('#d6472b')

    return f0, axess, axess1

def plot_coord_injpulse():
    """

    """

    f, ax= plt.subplots(nrows = 1, ncols = 2, num = 442365, figsize = (12,5), tight_layout = True)
    ax_0 = ax[0]
    ax_1 = ax[1]

    #injection pulse
    files_in_dir = os.listdir("./10")
    counter = 0
    for fil in files_in_dir:
        if fil[:2] == "HI":
            bunch10 = syttensen.bunch_syttensen.Bunch()
            bunch10.read_data("./10/%s" %fil)
        if counter == 0:
            big_bunch = bunch10
            counter+=1
            continue
        else:
            big_bunch.concat(bunch10)
        counter+=1
    big_bunch.plot()
    print(big_bunch.epoch)
    print((big_bunch.epoch[-1] - big_bunch.epoch[0])/3600)
    datetime_obj = []
    for j in np.arange(np.size(big_bunch.epoch)):
        datetime_obj_j = datetime.fromtimestamp(big_bunch.epoch[j])
        datetime_obj.append(datetime_obj_j)
    big_bunch.date_time = np.array((datetime_obj))
    print(big_bunch.date_time)


    ax_0_1 = ax_0.twinx()
    # axess.plot_date(big_bunch.date_time, big_bunch.h2o, linewidth  = 1, fmt = "-", color = "#332e2c")
    # axess1.plot_date(big_bunch.date_time, big_bunch.dD, linewidth = 1, fmt = "-", color = "#db724f")
    # date_format = mpl_dates.DateFormatter('%H:%M')
    # axess.xaxis.set_major_formatter(date_format)
    ax_0_1.plot(big_bunch.secs, big_bunch.h2o, linewidth  = 1, color = "#332e2c")
    ax_0.plot(big_bunch.secs, big_bunch.dD, linewidth = 1, color = "#d6472b")
    ax_0.set_xlabel("Run time [s]")
    ax_0.set_ylabel(r'$\delta\mathrm{D}$ [\textperthousand]', color = '#d6472b')
    ax_0_1.set_ylabel(r'$[{\mathrm{H}}_{2}\mathrm{O}] [\mathrm{ppm}]$', color = '#332e2c', rotation = 270, labelpad = 20)
    ax_0.set_ylim((-600, 0))
    ax_0.spines["left"].set_color('#d6472b')
    ax_0_1.spines["left"].set_color('#d6472b')
    ax_0_1.spines["right"].set_color('#332e2c')
    ax_0.spines["right"].set_color('#332e2c')
    ax_0.tick_params(axis = 'y', color = '#d6472b', labelcolor = '#d6472b')
    ax_0_1.tick_params(axis = 'y', color = '#332e2c', labelcolor = '#332e2c')
    ax_0.text(-0.1, -0.15, '(a)', transform = ax_0.transAxes, fontsize = 20)



    f = open("./HIDS2159_Fast_IsoWater_20160310_125444_corr.txt", "r")
    data_in = np.genfromtxt(f, skip_header = 1, usecols = (0,2,3,5,6,7), \
    dtype = None, names = ["nr", "dtime_string", "tray_pos", "d18", "dD", "h2o"], encoding = "None", delimiter = "\t", autostrip = True)
    print(data_in["dD"])
    print(data_in)
    dtime_obj = []
    for j in np.arange(np.size(data_in["dtime_string"])):
        dtime_obj.append(datetime.strptime(data_in["dtime_string"][j], "%Y/%m/%d %H:%M:%S"))
        print(type(dtime_obj[j]))
        data_in["tray_pos"][j] = np.int(data_in["tray_pos"][j][-2:])

    ax_1.plot_date(dtime_obj, data_in["dD"], linewidth  = 1.2, fmt = "-", color = "#d6472b")
    ax_1_1 = ax_1.twinx()
    ax_1_1.plot_date(dtime_obj, data_in["tray_pos"], linewidth = 0.78, fmt = "-", color = "#787675")
    date_format = mpl_dates.DateFormatter('%H:%M')
    ax_1.xaxis.set_major_formatter(date_format)
    ax_1_1.set_yticks([0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55])
    ax_1.set_xlabel("Time")
    ax_1.set_ylabel(r"$\delta\mathrm{D}$ [\textperthousand]", color = '#d6482b')
    ax_1_1.set_ylabel("Tray position", rotation = 270, labelpad = 20)
    ax_1.tick_params(axis = 'y', color = '#d6472b', labelcolor = '#d6472b')
    ax_1.spines["left"].set_color('#d6472b')
    ax_1_1.spines["left"].set_color('#d6472b')
    ax_1.text(-0.1, -0.15, '(b)', transform = ax_1.transAxes, fontsize = 20)




    return

# runs_series = return_unique_runs(instrument = "HKDS")
# print(np.size(runs_series))
# time.sleep(1)
# correct_standard_values_openpyxl(runs_series = runs_series)
# runs_build_data_frame_corr(runs_series, printout_df = True, save_csv = True)
# plot_offset_std2_vs_depth_vs_time()
# plot_std_std2_vs_depth_vs_time()
# plot_time_vs_depth()
# plot_nd_all_vs_instrument()
# plot_neem_nd_all()
# plot_offset_std_std2_histograms()
# plot_offset_std_std2_vs_depth()

# plot_injection_pulse()
# plot_cordinator_run()
plot_coord_injpulse()
